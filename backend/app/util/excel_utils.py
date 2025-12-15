"""
Excel 工具模块：不依赖任何模板文件，纯代码生成“示例模板那种产品级美观”的报表。

目标：
- 生成与示例模板同等级的观感：Dashboard + 数据源 + 多明细页
- 所有页统一主题：深色表头、表格条纹、边框、冻结窗格、打印设置、列宽、换行、条件格式
- Dashboard：KPI 卡片 + 状态分布图 + 分数分布图 + 维度均分图 + 低分 Top10
- 成绩总览：含“模型通过率(公式)” + “查看模型/查看细则(跳转)” + 条件格式（状态红绿、分数色阶）
- 模型结果（长表）：每学生每模型一行，便于筛选/透视
- 模型结果（宽表）：三模型横向对比
- 维度汇总：每学生一行，维度列动态生成（来自 sections.name），默认优先示例的 5 维
- 细则明细：主模型 items 扣分清单（扣分、原因说明、改进建议）
- 批次总览：KV 表
- 错误统计：类型/数量/占比(公式)
- 异常清单单独导出：同主题美观

输入 rows 结构沿用你原逻辑：
row = {
  "file_name": ...,
  "student_id": ...,
  "student_name": ...,
  "status": ...,
  "score": ...,
  "score_rubric": ...,
  "score_rubric_max": ...,
  "error_message": ...,
  "comment": ...,
  "aggregate_strategy": ...,
  "grader_results": [
     {
       "model_index": 1/2/3,
       "model_name": ...,
       "status": ...,
       "score": ...,
       "latency_ms": ...,
       "comment": ...,
       "error_message": ...,
       "sections": [
          {
            "name": "正确性",
            "score": ...,
            "max_score": ...,
            "comment": ...,
            "items": [
               {"name": "...", "score": ..., "max_score": ..., "comment": ..., "suggestion": ...}
            ]
          }
       ]
     }
  ]
}

注意：
- 本实现不使用任何外部 xlsx 模板文件。
- 如果你要把样式进一步“像素级对齐”某个示例模板，可在 _theme() / _layout 常量区微调色值与列宽。
"""

from __future__ import annotations

import math
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.util.logger import logger


def _规范化表头文本(text: Any) -> str:
    return str(text or "").replace(" ", "").replace("\u3000", "").strip()


class ExcelExporter:
    """
    纯代码生成（不依赖模板）的成绩与异常导出工具。
    """

    def __init__(self, batch_dir: Path) -> None:
        self.batch_dir = batch_dir

    # =========================================================
    # 主题与基础样式
    # =========================================================
    @staticmethod
    def _theme() -> dict[str, str]:
        # 颜色尽量贴近“企业蓝 + 浅灰底 + 柔和强调色”
        return {
            "navy": "0B2F5B",          # Dashboard 标题条
            "header": "1F4E79",        # 表头深蓝
            "header_font": "FFFFFF",
            "bg": "FFFFFF",
            "card": "F2F5F9",
            "card_border": "D6DCE5",
            "grid_border": "E0E0E0",
            "zebra": "F8FBFF",
            "meta": "D9E1F2",
            "score": "FFF2CC",
            "comment": "E2F0D9",
            "danger": "F8CBAD",
            "link": "0563C1",
            "text": "2B2B2B",
            "muted": "6B7280",
        }

    @staticmethod
    def _set_sheet_view(sheet, *, zoom: int = 110, show_grid: bool = False) -> None:
        sheet.sheet_view.zoomScale = zoom
        sheet.sheet_view.showGridLines = show_grid

    @staticmethod
    def _set_print(sheet, *, landscape: bool) -> None:
        sheet.print_title_rows = "1:1"
        sheet.page_setup.orientation = "landscape" if landscape else "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    @staticmethod
    def _border_thin(color: str) -> Border:
        s = Side(border_style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    # =========================================================
    # 表格（Excel Table）与通用排版
    # =========================================================
    @staticmethod
    def _apply_table(
        sheet,
        *,
        table_name: str,
        header_row: int,
        min_col: int,
        max_col: int,
        last_row: int,
        style: str = "TableStyleMedium9",
    ) -> Table:
        ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{last_row}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name=style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(tbl)
        return tbl

    def _style_header_row(self, sheet, *, header_row: int = 1, freeze: str = "A2") -> None:
        t = self._theme()
        fill = PatternFill("solid", fgColor=t["header"])
        font = Font(bold=True, color=t["header_font"])
        border = self._border_thin("DDDDDD")
        sheet.row_dimensions[header_row].height = 24
        for col in range(1, sheet.max_column + 1):
            c = sheet.cell(row=header_row, column=col)
            c.fill = fill
            c.font = font
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.freeze_panes = freeze

    def _apply_borders_all(self, sheet, *, start_row: int = 1) -> None:
        t = self._theme()
        border = self._border_thin(t["grid_border"])
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = border

    def _apply_zebra(self, sheet, *, start_row: int = 2) -> None:
        # Table 自带条纹，但某些查看器不渲染；这里用条件格式兜底
        t = self._theme()
        if sheet.max_row < start_row:
            return
        rng = f"A{start_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        sheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=["MOD(ROW(),2)=0"], fill=PatternFill("solid", fgColor=t["zebra"])),
        )

    @staticmethod
    def _wrap_column(sheet, col_letter: str, *, start_row: int = 2) -> None:
        for r in range(start_row, sheet.max_row + 1):
            c = sheet[f"{col_letter}{r}"]
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    @staticmethod
    def _left_top_align_all(sheet) -> None:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                old = cell.alignment
                wrap = bool(old.wrap_text) if old else False
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)

    @staticmethod
    def _auto_fit_columns(sheet, *, min_width: float = 10.0, max_width: float = 44.0, padding: float = 2.0) -> None:
        if sheet.max_column <= 0:
            return
        # 采样最多前 400 行（含表头）
        max_row = min(sheet.max_row, 400)
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            best = 0.0
            for r in range(1, max_row + 1):
                v = sheet.cell(r, col).value
                if v is None:
                    continue
                s = str(v)
                # 多行文本只取最长行
                if "\n" in s:
                    s = max(s.splitlines(), key=len) if s.splitlines() else s
                best = max(best, float(len(s)))
            width = min(max_width, max(min_width, best * 1.2 + padding))
            sheet.column_dimensions[letter].width = width

    @staticmethod
    def _cap_col(sheet, col_letter: str, max_width: float) -> None:
        dim = sheet.column_dimensions[col_letter]
        if dim.width is None:
            return
        dim.width = min(float(dim.width), float(max_width))

    @staticmethod
    def _min_col(sheet, col_letter: str, min_width: float) -> None:
        dim = sheet.column_dimensions[col_letter]
        cur = float(dim.width) if dim.width else 0.0
        dim.width = max(cur, float(min_width))

    @staticmethod
    def _set_number_format_col(sheet, col_index: int, fmt: str, *, start_row: int = 2) -> None:
        for r in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(r, col_index)
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt

    # =========================================================
    # 条件格式：状态红绿、分数色阶、关键列底色
    # =========================================================
    def _cf_status(self, sheet, col_letter: str, *, start_row: int = 2) -> None:
        if sheet.max_row < start_row:
            return
        t = self._theme()
        rng = f"{col_letter}{start_row}:{col_letter}{sheet.max_row}"
        sheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'EXACT(${col_letter}{start_row},"成功")'], fill=PatternFill("solid", fgColor=t["comment"])),
        )
        sheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'EXACT(${col_letter}{start_row},"失败")'], fill=PatternFill("solid", fgColor=t["danger"])),
        )

    def _cf_score_scale(self, sheet, col_letter: str, *, start_row: int = 2) -> None:
        if sheet.max_row < start_row:
            return
        # 红-黄-绿色阶
        rng = f"{col_letter}{start_row}:{col_letter}{sheet.max_row}"
        sheet.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    def _cf_static_fill(self, sheet, col_letter: str, fill_hex: str, *, start_row: int = 2) -> None:
        if sheet.max_row < start_row:
            return
        rng = f"{col_letter}{start_row}:{col_letter}{sheet.max_row}"
        sheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=["TRUE"], fill=PatternFill("solid", fgColor=fill_hex)),
        )

    # =========================================================
    # 业务：标准化与解析
    # =========================================================
    @staticmethod
    def _normalize_status(value: Any) -> str:
        text = str(value or "").strip()
        low = text.lower()
        if low in {"success", "成功"}:
            return "成功"
        if low in {"fail", "failed", "失败", "error", "异常"}:
            return "失败"
        return text

    @staticmethod
    def _detect_model_count(rows: list[dict]) -> int:
        max_idx = 1
        for row in rows:
            results = row.get("grader_results") or []
            if not isinstance(results, list):
                continue
            for item in results:
                if not isinstance(item, dict):
                    continue
                try:
                    idx = int(item.get("model_index") or 0)
                except Exception:
                    continue
                max_idx = max(max_idx, idx)
        return max(1, min(3, max_idx))

    @staticmethod
    def _model_label(idx: int) -> str:
        if idx == 1:
            return "主模型"
        if idx == 2:
            return "副模型A"
        if idx == 3:
            return "副模型B"
        return f"模型{idx}"

    def _normalize_model_comment(self, info: dict[str, Any]) -> str:
        if self._normalize_status(info.get("status")) != "成功":
            err = str(info.get("error_message") or "").strip()
            cmt = str(info.get("comment") or "").strip()
            return err or cmt or "失败"
        return str(info.get("comment") or "").strip()

    @staticmethod
    def _extract_sections(info: dict[str, Any]) -> list[dict[str, Any]]:
        sections = info.get("sections")
        if not isinstance(sections, list):
            return []
        out: list[dict[str, Any]] = []
        for sec in sections:
            if isinstance(sec, dict) and sec.get("name"):
                out.append(sec)
        return out

    # =========================================================
    # Dashboard：布局与图表（图表数据来自“数据”sheet 固定区域）
    # =========================================================
    def _build_dashboard(self, wb: Workbook, *, title: str) -> None:
        """
        创建 Dashboard 与 数据 两张表，并写固定布局。
        后续刷新数值只需要改“数据”与 Dashboard 的 KPI 单元格。
        """
        t = self._theme()

        # 1) Dashboard
        if "Dashboard" in wb.sheetnames:
            ws_dash = wb["Dashboard"]
            # 清空内容（保留 sheet）
            for row in ws_dash.iter_rows():
                for c in row:
                    c.value = None
        else:
            ws_dash = wb.create_sheet("Dashboard", 0)

        self._set_sheet_view(ws_dash, zoom=110, show_grid=False)
        self._set_print(ws_dash, landscape=True)

        # 列宽：A~L
        widths = {
            "A": 2,
            "B": 18,
            "C": 18,
            "D": 2,
            "E": 18,
            "F": 18,
            "G": 2,
            "H": 18,
            "I": 18,
            "J": 18,
            "K": 18,
            "L": 2,
        }
        for k, v in widths.items():
            ws_dash.column_dimensions[k].width = v

        # 顶部标题条
        ws_dash.merge_cells("A1:L1")
        c = ws_dash["A1"]
        c.value = title
        c.fill = PatternFill("solid", fgColor=t["navy"])
        c.font = Font(bold=True, size=18, color=t["header_font"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws_dash.row_dimensions[1].height = 34

        # 副标题（生成时间/总提交/成功率）
        ws_dash.merge_cells("A3:L3")
        ws_dash["A3"].value = "生成时间：-   |   总提交：-   |   成功率：-"
        ws_dash["A3"].font = Font(color=t["muted"])
        ws_dash["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws_dash.row_dimensions[3].height = 20

        # KPI 卡片（3 张大卡 + 3 张小卡）
        # 大卡：B5:C7、E5:F7、H5:K7
        def card_range(rng: str, title_text: str, value_cell: str) -> None:
            # 解析范围
            top_left = rng.split(":")[0]
            bottom_right = rng.split(":")[1]
            min_col = ord(top_left[0]) - ord("A") + 1
            min_row = int(top_left[1:])
            max_col = ord(bottom_right[0]) - ord("A") + 1
            max_row = int(bottom_right[1:])
            
            # 按行合并，而不是整个区域合并
            for rr in range(min_row, max_row + 1):
                start_cell = f"{chr(ord('A') + min_col - 1)}{rr}"
                end_cell = f"{chr(ord('A') + max_col - 1)}{rr}"
                ws_dash.merge_cells(f"{start_cell}:{end_cell}")
            
            # 设置标题
            cell = ws_dash[top_left]
            cell.value = f"{title_text}\n"
            cell.font = Font(bold=True, size=12, color=t["text"])
            cell.fill = PatternFill("solid", fgColor=t["card"])
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # 边框：对合并区域逐格加
            border = self._border_thin(t["card_border"])
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    ws_dash.cell(rr, cc).border = border
                    ws_dash.cell(rr, cc).fill = PatternFill("solid", fgColor=t["card"])

            # 数值单元格 - 使用第一列的对应行
            value_row = int(value_cell[1:])
            value_col_letter = chr(ord('A') + min_col - 1)
            actual_value_cell = f"{value_col_letter}{value_row}"
            ws_dash[actual_value_cell].value = "-"
            ws_dash[actual_value_cell].font = Font(bold=True, size=22, color=t["text"])
            ws_dash[actual_value_cell].alignment = Alignment(horizontal="left", vertical="center")

        # 大卡内容：
        # B5:C7 总提交
        card_range("B5:C7", "总提交", "B5")
        # E5:F7 成功/失败
        card_range("E5:F7", "成功 / 失败", "E5")
        # H5:K7 平均分 - 按行合并
        for rr in range(5, 8):
            ws_dash.merge_cells(f"H{rr}:K{rr}")
        ws_dash["H5"].value = "平均分\n"
        ws_dash["H5"].font = Font(bold=True, size=12, color=t["text"])
        ws_dash["H5"].fill = PatternFill("solid", fgColor=t["card"])
        ws_dash["H5"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for rr in range(5, 8):
            for cc in range(8, 12):
                ws_dash.cell(rr, cc).border = self._border_thin(t["card_border"])
                ws_dash.cell(rr, cc).fill = PatternFill("solid", fgColor=t["card"])
        ws_dash["H5"].value = "-"
        ws_dash["H5"].font = Font(bold=True, size=22, color=t["text"])
        ws_dash["H5"].alignment = Alignment(horizontal="left", vertical="center")
        ws_dash["E6"].value = "成功率 -"
        ws_dash["E6"].font = Font(color=t["muted"])
        ws_dash["H6"].value = "中位数 -"
        ws_dash["H6"].font = Font(color=t["muted"])

        # 小卡：B9:C10、E9:F10、H9:K10
        def small_card(rng: str, title_text: str, value_cell: str) -> None:
            # 解析范围
            top_left = rng.split(":")[0]
            bottom_right = rng.split(":")[1]
            min_col = ord(top_left[0]) - ord("A") + 1
            min_row = int(top_left[1:])
            max_col = ord(bottom_right[0]) - ord("A") + 1
            max_row = int(bottom_right[1:])
            
            # 按行合并
            for rr in range(min_row, max_row + 1):
                start_cell = f"{chr(ord('A') + min_col - 1)}{rr}"
                end_cell = f"{chr(ord('A') + max_col - 1)}{rr}"
                ws_dash.merge_cells(f"{start_cell}:{end_cell}")
            
            # 设置标题
            cell = ws_dash[top_left]
            cell.value = title_text
            cell.font = Font(bold=True, size=12, color=t["text"])
            cell.fill = PatternFill("solid", fgColor=t["card"])
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            border = self._border_thin(t["card_border"])
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    ws_dash.cell(rr, cc).border = border
                    ws_dash.cell(rr, cc).fill = PatternFill("solid", fgColor=t["card"])

            # 数值单元格 - 使用第一列的对应行
            value_row = int(value_cell[1:])
            value_col_letter = chr(ord('A') + min_col - 1)
            actual_value_cell = f"{value_col_letter}{value_row}"
            ws_dash[actual_value_cell].value = "-"
            ws_dash[actual_value_cell].font = Font(bold=True, size=14, color=t["text"])
            ws_dash[actual_value_cell].alignment = Alignment(horizontal="left", vertical="center")

        small_card("B9:C10", "P90", "B9")
        small_card("E9:F10", "最低 / 最高", "E9")
        # H9:K10 模型失败数 - 按行合并
        for rr in range(9, 11):
            ws_dash.merge_cells(f"H{rr}:K{rr}")
        ws_dash["H9"].value = "模型失败数"
        ws_dash["H9"].font = Font(bold=True, size=12, color=t["text"])
        ws_dash["H9"].fill = PatternFill("solid", fgColor=t["card"])
        ws_dash["H9"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for rr in range(9, 11):
            for cc in range(8, 12):
                ws_dash.cell(rr, cc).border = self._border_thin(t["card_border"])
                ws_dash.cell(rr, cc).fill = PatternFill("solid", fgColor=t["card"])
        ws_dash["H9"].value = "-"
        ws_dash["H9"].font = Font(bold=True, size=14, color=t["text"])
        ws_dash["H10"].value = "主模型 - | 副模型 -"
        ws_dash["H10"].font = Font(color=t["muted"])

        # 分区标题：图表与Top
        ws_dash["B12"].value = "状态分布"
        ws_dash["B12"].font = Font(bold=True, size=13, color=t["text"])
        ws_dash["E12"].value = "分数分布"
        ws_dash["E12"].font = Font(bold=True, size=13, color=t["text"])
        ws_dash["H12"].value = "维度均分"
        ws_dash["H12"].font = Font(bold=True, size=13, color=t["text"])
        ws_dash["B27"].value = "低分 Top10"
        ws_dash["B27"].font = Font(bold=True, size=13, color=t["text"])

        # 低分 Top 表头
        low_headers = ["文件名", "学号", "姓名", "最终分"]
        start_col = 2  # B
        header_row = 28
        for i, h in enumerate(low_headers):
            cell = ws_dash.cell(row=header_row, column=start_col + i, value=h)
            cell.fill = PatternFill("solid", fgColor=t["header"])
            cell.font = Font(bold=True, color=t["header_font"])
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = self._border_thin("DDDDDD")
        # 预留 10 行
        for r in range(header_row + 1, header_row + 11):
            for ccol in range(start_col, start_col + len(low_headers)):
                ws_dash.cell(r, ccol).border = self._border_thin(t["grid_border"])
                ws_dash.cell(r, ccol).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # 斑马
        rng = f"B{header_row+1}:E{header_row+10}"
        ws_dash.conditional_formatting.add(
            rng,
            FormulaRule(formula=["MOD(ROW(),2)=0"], fill=PatternFill("solid", fgColor=t["zebra"])),
        )

        # 2) 数据 sheet：作为 Dashboard 图表数据源
        if "数据" in wb.sheetnames:
            ws_data = wb["数据"]
            for row in ws_data.iter_rows():
                for c in row:
                    c.value = None
        else:
            ws_data = wb.create_sheet("数据", 1)

        self._set_sheet_view(ws_data, zoom=100, show_grid=False)
        self._set_print(ws_data, landscape=False)

        # 数据区布局（固定）
        # 状态分布 A1:B3
        ws_data["A1"].value = "状态"
        ws_data["B1"].value = "数量"
        ws_data["A2"].value = "成功"
        ws_data["A3"].value = "失败"
        ws_data["B2"].value = 0
        ws_data["B3"].value = 0

        # 分数分布 A5:B15（10 档）
        ws_data["A5"].value = "分数段"
        ws_data["B5"].value = "人数"
        for i in range(10):
            ws_data[f"A{6+i}"].value = f"{i*10:02d}-{i*10+9:02d}"
            ws_data[f"B{6+i}"].value = 0

        # 维度均分 D1:E20（最多 18 个维度，够用；超出再扩）
        ws_data["D1"].value = "维度"
        ws_data["E1"].value = "均分"
        for i in range(2, 20):
            ws_data[f"D{i}"].value = ""
            ws_data[f"E{i}"].value = ""

        # 给数据表头一点基础样式（虽不展示，但干净）
        for (a, b) in (("A1", "B1"), ("A5", "B5"), ("D1", "E1")):
            ws_data[a].font = Font(bold=True)
            ws_data[b].font = Font(bold=True)

        # 3) Dashboard 图表（引用 数据 sheet 固定范围）
        # 状态分布：饼图
        pie = PieChart()
        pie.title = "状态分布"
        data = Reference(ws_data, min_col=2, min_row=1, max_row=3)   # B1:B3
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=3)   # A2:A3
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.height = 6.0
        pie.width = 9.0
        ws_dash.add_chart(pie, "B13")

        # 分数分布：柱状图
        bar = BarChart()
        bar.type = "col"
        bar.title = "分数分布"
        bar.y_axis.title = "人数"
        bar.x_axis.title = "分数段"
        data2 = Reference(ws_data, min_col=2, min_row=5, max_row=15)  # B5:B15
        cats2 = Reference(ws_data, min_col=1, min_row=6, max_row=15)  # A6:A15
        bar.add_data(data2, titles_from_data=True)
        bar.set_categories(cats2)
        bar.height = 6.0
        bar.width = 12.0
        ws_dash.add_chart(bar, "E13")

        # 维度均分：柱状图（D1:E19）
        bar2 = BarChart()
        bar2.type = "col"
        bar2.title = "维度均分"
        bar2.y_axis.title = "均分"
        bar2.x_axis.title = "维度"
        data3 = Reference(ws_data, min_col=5, min_row=1, max_row=19)  # E1:E19
        cats3 = Reference(ws_data, min_col=4, min_row=2, max_row=19)  # D2:D19
        bar2.add_data(data3, titles_from_data=True)
        bar2.set_categories(cats3)
        bar2.height = 6.0
        bar2.width = 12.0
        ws_dash.add_chart(bar2, "H13")

    # =========================================================
    # Sheet 创建：通用“带表格”的页
    # =========================================================
    def _create_table_sheet(
        self,
        wb: Workbook,
        *,
        title: str,
        headers: list[str],
        table_name: str,
        landscape: bool,
        freeze: str,
    ):
        if title in wb.sheetnames:
            ws = wb[title]
            # 清空 sheet（保留对象）
            for row in ws.iter_rows():
                for c in row:
                    c.value = None
            # 清空 tables
            ws.tables.clear()
        else:
            ws = wb.create_sheet(title)

        self._set_sheet_view(ws, zoom=110, show_grid=False)
        self._set_print(ws, landscape=landscape)

        ws.append(headers)
        self._style_header_row(ws, header_row=1, freeze=freeze)

        # 先放一个空数据行占位（Table 要有至少一行）
        ws.append([None] * len(headers))

        tbl = self._apply_table(
            ws,
            table_name=table_name,
            header_row=1,
            min_col=1,
            max_col=len(headers),
            last_row=2,
            style="TableStyleMedium9",
        )

        return ws, tbl

    # =========================================================
    # 刷新数据：把统计写入 数据/Dashboard
    # =========================================================
    def _refresh_dashboard_numbers(
        self,
        wb: Workbook,
        *,
        batch_title: str,
        total: int,
        ok: int,
        fail: int,
        scores_sorted: list[float],
        dim_avg: dict[str, float],
        model_fail_by_idx: dict[int, int],
        low_top: list[tuple[Any, Any, Any, Any]],
    ) -> None:
        if "Dashboard" not in wb.sheetnames or "数据" not in wb.sheetnames:
            return
        t = self._theme()
        ws_dash = wb["Dashboard"]
        ws_data = wb["数据"]

        ok_rate = f"{(ok / total * 100):.1f}%" if total else "0.0%"

        if scores_sorted:
            mean = round(float(statistics.mean(scores_sorted)), 1)
            median = round(float(statistics.median(scores_sorted)), 1)
            p90_idx = max(0, min(len(scores_sorted) - 1, int(math.ceil(len(scores_sorted) * 0.9)) - 1))
            p90 = round(float(scores_sorted[p90_idx]), 1)
            mn = round(float(scores_sorted[0]), 1)
            mx = round(float(scores_sorted[-1]), 1)
        else:
            mean = None
            median = None
            p90 = None
            mn = None
            mx = None

        # 标题与副标题
        ws_dash["A1"].value = batch_title
        ws_dash["A3"].value = (
            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}   |   总提交：{total}   |   成功率：{ok_rate}"
        )
        ws_dash["A3"].font = Font(color=t["muted"])

        # KPI (使用合并单元格的左上角单元格)
        ws_dash["B5"].value = total
        ws_dash["E5"].value = f"{ok} / {fail}"
        ws_dash["H5"].value = mean if mean is not None else "-"
        ws_dash["E6"].value = f"成功率 {ok_rate}"
        ws_dash["H6"].value = f"中位数 {median:.1f}" if median is not None else "中位数 -"
        ws_dash["B9"].value = p90 if p90 is not None else "-"
        ws_dash["E9"].value = f"{mn:.1f} / {mx:.1f}" if (mn is not None and mx is not None) else "- / -"

        total_model_fail = sum(model_fail_by_idx.values())
        ws_dash["H9"].value = total_model_fail
        ws_dash["H10"].value = (
            f"主模型 {model_fail_by_idx.get(1, 0)} | 副模型 {model_fail_by_idx.get(2, 0) + model_fail_by_idx.get(3, 0)}"
        )

        # 低分 Top10
        start_row = 29
        for i in range(10):
            r = start_row + i
            if i < len(low_top):
                fn, sid, sname, sc = low_top[i]
                ws_dash.cell(r, 2).value = fn
                ws_dash.cell(r, 3).value = sid
                ws_dash.cell(r, 4).value = sname
                ws_dash.cell(r, 5).value = sc
            else:
                ws_dash.cell(r, 2).value = None
                ws_dash.cell(r, 3).value = None
                ws_dash.cell(r, 4).value = None
                ws_dash.cell(r, 5).value = None

        # 数据 sheet：状态
        ws_data["B2"].value = ok
        ws_data["B3"].value = fail

        # 数据 sheet：分数桶（10 档）
        bins_counts = [0] * 10
        for s in scores_sorted:
            try:
                x = float(s)
            except Exception:
                continue
            idx = int(x // 10)
            idx = max(0, min(9, idx))
            bins_counts[idx] += 1
        for i in range(10):
            ws_data[f"B{6+i}"].value = bins_counts[i]

        # 数据 sheet：维度均分（最多 18 个维度）
        # 清空旧
        for r in range(2, 20):
            ws_data[f"D{r}"].value = ""
            ws_data[f"E{r}"].value = ""
        # 填新（按均分降序展示，观感更像“看板”）
        items = sorted(dim_avg.items(), key=lambda kv: (-float(kv[1]), str(kv[0])))
        for i, (k, v) in enumerate(items[:18], start=2):
            ws_data[f"D{i}"].value = k
            ws_data[f"E{i}"].value = round(float(v), 1)

    # =========================================================
    # 导出：成绩报表（主入口）
    # =========================================================
    def export_results(
        self,
        rows: Iterable[dict],
        summary: Optional[dict[str, Any]] = None,
        error_rows: Optional[Iterable[dict]] = None,
    ) -> Path:
        rows_list = list(rows)
        model_count = self._detect_model_count(rows_list)

        # 1) 建 workbook 与 Dashboard/数据
        wb = Workbook()
        # openpyxl 默认会建一个 Sheet，删掉
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        self._build_dashboard(wb, title=f"作业批改结果 · {self.batch_dir.name}")

        # 2) 业务页：创建并写表头
        # 成绩总览（注意：列顺序更接近“产品表”，易读且像示例）
        overview_headers = [
            "状态",
            "最终分",
            "姓名",
            "学号",
            "文件名",
            "成功模型数",
            "失败模型数",
            "模型通过率",
            "总体评语",
            "错误描述",
            "聚合算法",
            "查看模型",
            "查看细则",
        ]
        ws_overview, tbl_overview = self._create_table_sheet(
            wb,
            title="成绩总览",
            headers=overview_headers,
            table_name="OverviewTbl",
            landscape=False,
            freeze="A2",
        )

        # 模型结果（长表）
        model_long_headers = [
            "学号",
            "姓名",
            "文件名",
            "模型",
            "是否采用",
            "状态",
            "分数",
            "耗时(ms)",
            "评语",
        ]
        ws_mlong, tbl_mlong = self._create_table_sheet(
            wb,
            title="模型结果（长表）",
            headers=model_long_headers,
            table_name="ModelLong",
            landscape=True,
            freeze="A2",
        )

        # 模型结果（宽表）
        wide_headers = ["学号", "姓名", "文件名"]
        for idx in range(1, model_count + 1):
            label = self._model_label(idx)
            wide_headers += [
                f"{label}-状态",
                f"{label}-分数",
                f"{label}-耗时(ms)",
                f"{label}-评语",
            ]
        wide_headers += ["总体评语"]
        ws_mwide, tbl_mwide = self._create_table_sheet(
            wb,
            title="批改模型结果（宽表）",
            headers=wide_headers,
            table_name="ModelWide",
            landscape=True,
            freeze="A2",
        )

        # 维度汇总：维度列动态，但优先“示例常见 5 维”
        preferred_dims = ["正确性", "鲁棒性", "效率", "可读性", "规范性"]
        # 先扫描所有维度
        all_dims: set[str] = set(preferred_dims)
        for row in rows_list:
            results = row.get("grader_results") or []
            if not isinstance(results, list):
                continue
            for it in results:
                if not isinstance(it, dict):
                    continue
                for sec in self._extract_sections(it):
                    all_dims.add(str(sec.get("name") or "").strip())
        # 排序：优先 5 维在前，其余按字典序
        dims_ordered = preferred_dims + sorted([d for d in all_dims if d and d not in preferred_dims])

        dim_headers = ["学号", "姓名", "最终分"] + dims_ordered
        ws_dim, tbl_dim = self._create_table_sheet(
            wb,
            title="维度汇总",
            headers=dim_headers,
            table_name="DimSummary",
            landscape=True,
            freeze="A2",
        )

        # 细则明细（主模型 items 扣分清单）
        rubric_headers = ["学号", "姓名", "维度", "细则项", "扣分", "原因说明", "改进建议"]
        ws_rubric, tbl_rubric = self._create_table_sheet(
            wb,
            title="细则明细",
            headers=rubric_headers,
            table_name="RubricDetail",
            landscape=True,
            freeze="A2",
        )

        # 批次总览（KV）
        summary_headers = ["字段", "值"]
        ws_sum, tbl_sum = self._create_table_sheet(
            wb,
            title="批次总览",
            headers=summary_headers,
            table_name="BatchSummary",
            landscape=False,
            freeze="A2",
        )

        # 错误统计
        err_headers = ["错误类型", "数量", "占比"]
        ws_err, tbl_err = self._create_table_sheet(
            wb,
            title="错误统计",
            headers=err_headers,
            table_name="ErrStats",
            landscape=False,
            freeze="A2",
        )

        # 3) 清掉各表的占位空行（第 2 行）以便重新写真实数据
        def clear_placeholder(ws) -> None:
            # 第 2 行数据占位清空（值清空即可）
            for c in range(1, ws.max_column + 1):
                ws.cell(2, c).value = None

        for w in (ws_overview, ws_mlong, ws_mwide, ws_dim, ws_rubric, ws_sum, ws_err):
            clear_placeholder(w)

        # 4) 批次总览写入
        if summary:
            for k, v in summary.items():
                ws_sum.append([str(k), "" if v is None else str(v)])

        # 5) 错误统计写入（来自 error_rows 或 rows 里的失败信息聚合）
        err_counter = Counter()
        if error_rows:
            for e in error_rows:
                et = str(e.get("error_type") or "").strip() or "未知"
                err_counter[et] += 1
        else:
            for r in rows_list:
                if self._normalize_status(r.get("status")) == "失败":
                    et = str(r.get("error_type") or "失败").strip() or "失败"
                    err_counter[et] += 1
        err_total = sum(err_counter.values())
        for et, cnt in sorted(err_counter.items(), key=lambda kv: (-kv[1], kv[0])):
            ws_err.append([et, cnt, None])  # 占比后面补公式

        # 6) 主体写入：成绩总览 / 模型长表 / 模型宽表 / 维度 / 细则
        # 记录跳转行号
        first_row_model_long: dict[Any, int] = {}
        first_row_rubric: dict[Any, int] = {}

        # Dashboard 统计
        scores_for_stats: list[float] = []
        ok = 0
        fail = 0
        model_fail_by_idx: dict[int, int] = {1: 0, 2: 0, 3: 0}
        dim_sum_cnt: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0])  # dim -> [sum, cnt]
        low_candidates: list[tuple[Any, Any, Any, Any]] = []

        # 预计算一些列号（用于公式/超链接）
        # 成绩总览列：
        col_success = overview_headers.index("成功模型数") + 1
        col_fail = overview_headers.index("失败模型数") + 1
        col_pass = overview_headers.index("模型通过率") + 1
        col_link_model = overview_headers.index("查看模型") + 1
        col_link_rubric = overview_headers.index("查看细则") + 1
        col_student_id = overview_headers.index("学号") + 1

        for row in rows_list:
            file_name = row.get("file_name")
            student_id = row.get("student_id")
            student_name = row.get("student_name")
            status = self._normalize_status(row.get("status"))
            score = row.get("score")
            comment = row.get("comment")
            err_msg = row.get("error_message")
            aggregate_strategy = row.get("aggregate_strategy") or ""
            grader_results = row.get("grader_results") or []

            student_key = student_id if str(student_id or "").strip() else file_name

            # 统计
            if status == "成功":
                ok += 1
            elif status:
                fail += 1
            if isinstance(score, (int, float)):
                scores_for_stats.append(float(score))
                low_candidates.append((file_name, student_id, student_name, float(score)))

            # by_index
            by_index: dict[int, dict[str, Any]] = {}
            if isinstance(grader_results, list):
                for item in grader_results:
                    if isinstance(item, dict):
                        try:
                            idx = int(item.get("model_index"))
                        except Exception:
                            continue
                        if 1 <= idx <= 3:
                            by_index[idx] = item

            # 成功/失败模型数
            model_success_count = 0
            model_fail_count = 0
            for idx in range(1, model_count + 1):
                info = by_index.get(idx) or {}
                s = self._normalize_status(info.get("status"))
                if s == "成功":
                    model_success_count += 1
                elif s:
                    model_fail_count += 1
                    model_fail_by_idx[idx] = model_fail_by_idx.get(idx, 0) + 1

            # 成绩总览：先 append 一行，稍后补公式和超链接
            ws_overview.append(
                [
                    status,
                    score,
                    student_name,
                    student_id,
                    file_name,
                    model_success_count,
                    model_fail_count,
                    None,  # 通过率公式
                    comment,
                    err_msg,
                    aggregate_strategy,
                    None,  # 查看模型
                    None,  # 查看细则
                ]
            )
            overview_row_idx = ws_overview.max_row

            # 通过率公式
            s_col = get_column_letter(col_success)
            f_col = get_column_letter(col_fail)
            p_col = get_column_letter(col_pass)
            ws_overview[f"{p_col}{overview_row_idx}"].value = (
                f'=IF(({s_col}{overview_row_idx}+{f_col}{overview_row_idx})=0,"",'
                f'{s_col}{overview_row_idx}/({s_col}{overview_row_idx}+{f_col}{overview_row_idx}))'
            )
            ws_overview[f"{p_col}{overview_row_idx}"].number_format = "0.0%"

            # 模型宽表：对比各模型
            wide_row = [student_id, student_name, file_name]
            for idx in range(1, model_count + 1):
                info = by_index.get(idx) or {}
                wide_row += [
                    self._normalize_status(info.get("status")),
                    info.get("score"),
                    info.get("latency_ms"),
                    self._normalize_model_comment(info),
                ]
            wide_row += [comment]
            ws_mwide.append(wide_row)

            # 模型长表：每模型一行
            for idx in range(1, model_count + 1):
                info = by_index.get(idx) or {}
                if student_key not in first_row_model_long:
                    first_row_model_long[student_key] = ws_mlong.max_row + 1
                ws_mlong.append(
                    [
                        student_id,
                        student_name,
                        file_name,
                        self._model_label(idx),
                        "是" if idx == 1 else "",
                        self._normalize_status(info.get("status")),
                        info.get("score"),
                        info.get("latency_ms"),
                        self._normalize_model_comment(info),
                    ]
                )

            # 维度汇总：取“成功模型”的维度分数均值
            # 先组织各模型 sections
            sections_by_model: dict[int, list[dict[str, Any]]] = {}
            for idx in (1, 2, 3):
                sections_by_model[idx] = self._extract_sections(by_index.get(idx) or {})

            def mean_dim_for_student(dim_name: str) -> Optional[float]:
                vals: list[float] = []
                for idx in (1, 2, 3):
                    info = by_index.get(idx) or {}
                    if self._normalize_status(info.get("status")) != "成功":
                        continue
                    sec = next((s for s in sections_by_model[idx] if str(s.get("name") or "") == dim_name), None)
                    if isinstance(sec, dict) and isinstance(sec.get("score"), (int, float)):
                        vals.append(float(sec["score"]))
                if not vals:
                    return None
                m = float(statistics.mean(vals))
                acc = dim_sum_cnt[dim_name]
                acc[0] += m
                acc[1] += 1.0
                return round(m, 2)

            dim_values = [student_id, student_name, score]
            for d in dims_ordered:
                dim_values.append(mean_dim_for_student(d))
            ws_dim.append(dim_values)

            # 细则明细：只用主模型（idx=1）items 做扣分清单
            main = by_index.get(1) or {}
            main_sections = self._extract_sections(main)
            for sec in main_sections:
                dim_name = str(sec.get("name") or "").strip()
                items = sec.get("items") or []
                if not isinstance(items, list):
                    continue
                for it in items:
                    if not isinstance(it, dict) or not it.get("name"):
                        continue
                    if student_key not in first_row_rubric:
                        first_row_rubric[student_key] = ws_rubric.max_row + 1
                    item_name = str(it.get("name") or "").strip()

                    item_score = it.get("score")
                    item_max = it.get("max_score")
                    deduct: Optional[float] = None
                    if isinstance(item_score, (int, float)) and isinstance(item_max, (int, float)):
                        d = max(0.0, float(item_max) - float(item_score))
                        deduct = float(int(round(d))) if abs(d - round(d)) < 1e-6 else round(d, 2)

                    reason = str(it.get("comment") or it.get("reason") or it.get("evidence") or "").strip()
                    suggest = str(
                        it.get("suggestion")
                        or it.get("advice")
                        or it.get("improvement")
                        or it.get("recommendation")
                        or ""
                    ).strip()

                    ws_rubric.append([student_id, student_name, dim_name, item_name, deduct, reason, suggest])

            # 成绩总览：超链接（查看模型/查看细则）
            mrow = first_row_model_long.get(student_key)
            rrow = first_row_rubric.get(student_key)
            if mrow:
                cell = ws_overview.cell(row=overview_row_idx, column=col_link_model)
                cell.value = f'=HYPERLINK("#\'模型结果（长表）\'!A{mrow}","查看")'
                cell.font = Font(color=self._theme()["link"], underline="single")
            if rrow:
                cell = ws_overview.cell(row=overview_row_idx, column=col_link_rubric)
                cell.value = f'=HYPERLINK("#\'细则明细\'!A{rrow}","查看")'
                cell.font = Font(color=self._theme()["link"], underline="single")

        # 7) 错误统计占比公式（在写完行数后补）
        # ErrStats: B 列数量，C 列占比
        if ws_err.max_row >= 2:
            start = 2
            end = ws_err.max_row
            for r in range(start, end + 1):
                ws_err[f"C{r}"].value = f"=B{r}/SUM($B${start}:$B${end})"
                ws_err[f"C{r}"].number_format = "0.0%"

        # 8) 更新所有 Table ref（必须在数据写完后）
        def resize_table(ws, tbl: Table) -> None:
            # 若没有数据行，至少保留 2 行（header + 1 空行）
            last = max(2, ws.max_row)
            tbl.ref = f"A1:{get_column_letter(ws.max_column)}{last}"

        for ws, tbl in (
            (ws_overview, tbl_overview),
            (ws_mlong, tbl_mlong),
            (ws_mwide, tbl_mwide),
            (ws_dim, tbl_dim),
            (ws_rubric, tbl_rubric),
            (ws_sum, tbl_sum),
            (ws_err, tbl_err),
        ):
            resize_table(ws, tbl)

        # 9) 统一美观：边框、斑马、列宽、换行、条件格式、数字格式
        t = self._theme()

        def polish_table_sheet(ws, *, landscape: bool) -> None:
            self._set_sheet_view(ws, zoom=110, show_grid=False)
            self._set_print(ws, landscape=landscape)
            self._apply_borders_all(ws, start_row=1)
            self._apply_zebra(ws, start_row=2)
            self._left_top_align_all(ws)
            self._auto_fit_columns(ws, min_width=10, max_width=44, padding=2.0)

        polish_table_sheet(ws_overview, landscape=False)
        polish_table_sheet(ws_mlong, landscape=True)
        polish_table_sheet(ws_mwide, landscape=True)
        polish_table_sheet(ws_dim, landscape=True)
        polish_table_sheet(ws_rubric, landscape=True)
        polish_table_sheet(ws_sum, landscape=False)
        polish_table_sheet(ws_err, landscape=False)

        # 关键列：最小宽度/最大宽度 + 换行
        # 文件名列一般很长：限制宽度 + 开启换行
        # 成绩总览：文件名=E
        self._cap_col(ws_overview, "E", 30)
        self._wrap_column(ws_overview, "E", start_row=2)
        # 总体评语/错误描述：I/J
        self._min_col(ws_overview, "I", 28)
        self._min_col(ws_overview, "J", 28)
        self._wrap_column(ws_overview, "I", start_row=2)
        self._wrap_column(ws_overview, "J", start_row=2)

        # 模型长表：文件名 C、评语 I
        self._cap_col(ws_mlong, "C", 30)
        self._wrap_column(ws_mlong, "C", start_row=2)
        self._min_col(ws_mlong, "I", 32)
        self._wrap_column(ws_mlong, "I", start_row=2)

        # 宽表：文件名 C、各评语列 + 总体评语最后一列
        self._cap_col(ws_mwide, "C", 30)
        self._wrap_column(ws_mwide, "C", start_row=2)
        # 评语列：从 D 开始每 4 列一个模型，评语在第 4 列：G/K/O，再加最后列（总体评语）
        # 计算一下最后列
        last_col_letter = get_column_letter(ws_mwide.max_column)
        for col_letter in ("G", "K", "O", last_col_letter):
            if col_letter <= last_col_letter:
                self._min_col(ws_mwide, col_letter, 30)
                self._wrap_column(ws_mwide, col_letter, start_row=2)

        # 维度汇总：维度列多，限制最大宽度让表更紧凑
        for c in range(1, ws_dim.max_column + 1):
            letter = get_column_letter(c)
            self._cap_col(ws_dim, letter, 18)
        self._cap_col(ws_dim, "A", 16)
        self._cap_col(ws_dim, "B", 16)

        # 细则明细：原因/建议需要宽且换行
        self._min_col(ws_rubric, "F", 30)
        self._min_col(ws_rubric, "G", 30)
        self._wrap_column(ws_rubric, "F", start_row=2)
        self._wrap_column(ws_rubric, "G", start_row=2)
        self._cap_col(ws_rubric, "D", 26)  # 细则项
        self._wrap_column(ws_rubric, "D", start_row=2)

        # 批次总览：值列宽
        self._min_col(ws_sum, "A", 22)
        self._min_col(ws_sum, "B", 56)
        self._wrap_column(ws_sum, "B", start_row=2)

        # 错误统计：错误类型宽一点
        self._min_col(ws_err, "A", 22)

        # 数字格式：分数、耗时、扣分
        # 成绩总览：最终分=B
        self._set_number_format_col(ws_overview, 2, "0.00", start_row=2)
        # 模型长表：分数=G、耗时=H
        self._set_number_format_col(ws_mlong, 7, "0.00", start_row=2)
        self._set_number_format_col(ws_mlong, 8, "0", start_row=2)
        # 宽表：分数列（E/I/M）和耗时列（F/J/N）
        # 宽表结构：A 学号 B 姓名 C 文件名
        # 主模型：D 状态 E 分数 F 耗时 G 评语
        # 副A：H 状态 I 分数 J 耗时 K 评语
        # 副B：L 状态 M 分数 N 耗时 O 评语
        # 最后：总体评语 P（若存在）
        self._set_number_format_col(ws_mwide, 5, "0.00", start_row=2)
        self._set_number_format_col(ws_mwide, 6, "0", start_row=2)
        self._set_number_format_col(ws_mwide, 9, "0.00", start_row=2)
        self._set_number_format_col(ws_mwide, 10, "0", start_row=2)
        self._set_number_format_col(ws_mwide, 13, "0.00", start_row=2)
        self._set_number_format_col(ws_mwide, 14, "0", start_row=2)
        # 维度汇总：最终分=C 维度列 D.. 统一 0.00
        self._set_number_format_col(ws_dim, 3, "0.00", start_row=2)
        for c in range(4, ws_dim.max_column + 1):
            self._set_number_format_col(ws_dim, c, "0.00", start_row=2)
        # 细则：扣分=E
        self._set_number_format_col(ws_rubric, 5, "0.00", start_row=2)

        # 条件格式：成绩总览 状态列 A、最终分 B、通过率 H
        self._cf_status(ws_overview, "A", start_row=2)
        self._cf_score_scale(ws_overview, "B", start_row=2)
        self._cf_static_fill(ws_overview, "E", t["meta"], start_row=2)       # 文件名背景
        self._cf_static_fill(ws_overview, "I", t["comment"], start_row=2)    # 总体评语
        self._cf_static_fill(ws_overview, "J", t["comment"], start_row=2)    # 错误描述
        self._cf_static_fill(ws_overview, "B", t["score"], start_row=2)      # 最终分底色（叠加色阶也 OK）

        # 模型长表：状态 F、分数 G、评语 I
        self._cf_status(ws_mlong, "F", start_row=2)
        self._cf_score_scale(ws_mlong, "G", start_row=2)
        self._cf_static_fill(ws_mlong, "I", t["comment"], start_row=2)

        # 宽表：三模型状态/分数/评语
        for col in ("D", "H", "L"):
            self._cf_status(ws_mwide, col, start_row=2)
        for col in ("E", "I", "M"):
            self._cf_score_scale(ws_mwide, col, start_row=2)
        for col in ("G", "K", "O"):
            self._cf_static_fill(ws_mwide, col, t["comment"], start_row=2)

        # 错误统计：数量列底色
        self._cf_static_fill(ws_err, "B", t["meta"], start_row=2)

        # 10) 刷新 Dashboard 数据
        scores_sorted = sorted(scores_for_stats)
        low_top = sorted(low_candidates, key=lambda x: x[3])[:10]
        dim_avg: dict[str, float] = {}
        for k, (s, c) in dim_sum_cnt.items():
            if c:
                dim_avg[k] = float(s) / float(c)

        self._refresh_dashboard_numbers(
            wb,
            batch_title=f"作业批改结果 · {self.batch_dir.name}",
            total=len(rows_list),
            ok=ok,
            fail=fail,
            scores_sorted=scores_sorted,
            dim_avg=dim_avg,
            model_fail_by_idx=model_fail_by_idx,
            low_top=low_top,
        )

        # 11) 保存
        target = self.batch_dir / "grade_result.xlsx"
        wb.save(target)
        logger.info("成绩表已生成（纯代码，无模板）：%s", target)
        return target

    # =========================================================
    # 导出：异常清单（单表，纯代码美观）
    # =========================================================
    def export_errors(self, rows: Iterable[dict]) -> Path:
        rows_list = list(rows)

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet("异常清单", 0)

        self._set_sheet_view(ws, zoom=110, show_grid=False)
        self._set_print(ws, landscape=False)

        headers = ["文件名", "错误类型", "错误描述"]
        ws.append(headers)
        self._style_header_row(ws, header_row=1, freeze="A2")

        for r in rows_list:
            ws.append([r.get("file_name"), r.get("error_type"), r.get("error_message")])

        # Table
        last_row = max(2, ws.max_row)
        self._apply_table(
            ws,
            table_name="TblErrorList",
            header_row=1,
            min_col=1,
            max_col=len(headers),
            last_row=last_row,
            style="TableStyleMedium9",
        )

        # 美化
        self._apply_borders_all(ws, start_row=1)
        self._apply_zebra(ws, start_row=2)
        self._left_top_align_all(ws)

        # 列宽与换行
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 56
        self._wrap_column(ws, "C", start_row=2)

        # 条件格式：错误类型/描述淡色
        t = self._theme()
        self._cf_static_fill(ws, "A", t["meta"], start_row=2)
        self._cf_static_fill(ws, "B", t["danger"], start_row=2)
        self._cf_static_fill(ws, "C", t["comment"], start_row=2)

        target = self.batch_dir / "error_list.xlsx"
        wb.save(target)
        logger.info("异常清单已生成（纯代码，无模板）：%s", target)
        return target
