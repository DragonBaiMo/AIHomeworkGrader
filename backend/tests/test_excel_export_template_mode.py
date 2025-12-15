"""Excel 导出（模板法）覆盖测试。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.util.excel_utils import ExcelExporter


def _添加简单表格(ws, *, table_name: str, headers: list[str], sample_row: list[object]) -> None:
    ws.append(headers)
    ws.append(sample_row)
    ref = f"A1:{chr(ord('A') + len(headers) - 1)}2"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)


def test_export_results_模板法_保留表格并更新ref与超链接(tmp_path: Path) -> None:
    template_path = tmp_path / "grade_template.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    ws_overview = wb.create_sheet("成绩总览")
    _添加简单表格(
        ws_overview,
        table_name="OverviewTbl",
        headers=["文件名", "学号", "姓名", "状态", "最终分", "总体评语", "错误描述", "查看模型", "查看细则"],
        sample_row=["示例.docx", "000", "示例", "成功", 60, "示例评语", None, "", ""],
    )

    ws_models_long = wb.create_sheet("模型结果（长表）")
    _添加简单表格(
        ws_models_long,
        table_name="ModelLong",
        headers=["文件名", "学号", "姓名", "模型", "状态", "分数", "评语", "耗时(ms)"],
        sample_row=["示例.docx", "000", "示例", "主模型", "成功", 60, "示例评语", 1],
    )

    ws_models_wide = wb.create_sheet("批改模型结果（宽表）")
    _添加简单表格(
        ws_models_wide,
        table_name="ModelWide",
        headers=["文件名", "学号", "姓名", "主模型名称", "主模型状态", "主模型分数", "主模型评语", "主模型耗时(ms)", "LLM总评语"],
        sample_row=["示例.docx", "000", "示例", "main", "成功", 60, "示例评语", 1, "示例总评语"],
    )

    ws_dim = wb.create_sheet("维度汇总")
    _添加简单表格(
        ws_dim,
        table_name="DimSummary",
        headers=["文件名", "学号", "姓名", "维度名称", "汇总得分", "满分", "主模型得分", "分歧度", "维度评语"],
        sample_row=["示例.docx", "000", "示例", "结构", 8, 10, 8, 0, "示例评语"],
    )

    ws_criteria = wb.create_sheet("细则明细")
    _添加简单表格(
        ws_criteria,
        table_name="RubricDetail",
        headers=["文件名", "学号", "姓名", "维度名称", "细则名称", "汇总得分", "满分", "主模型得分", "分歧度", "扣分原因"],
        sample_row=["示例.docx", "000", "示例", "结构", "标题", 4, 5, 4, 0, "示例原因"],
    )

    ws_summary = wb.create_sheet("批次总览")
    _添加简单表格(ws_summary, table_name="BatchSummary", headers=["字段", "值"], sample_row=["批次ID", "demo"])

    wb.save(template_path)

    exporter = ExcelExporter(tmp_path, template_path=template_path)
    exporter.export_results(
        [
            {
                "file_name": "张三.docx",
                "student_id": "2025001",
                "student_name": "张三",
                "score": 55.0,
                "status": "成功",
                "comment": "总体评语",
                "error_message": None,
                "grader_results": [
                    {
                        "model_index": 1,
                        "model_name": "main",
                        "status": "成功",
                        "score": 55,
                        "comment": "c1",
                        "latency_ms": 100,
                        "sections": [
                            {
                                "name": "结构",
                                "max_score": 10,
                                "score": 8,
                                "comment": "结构不错",
                                "items": [
                                    {"name": "标题", "max_score": 5, "score": 4, "comment": "标题清晰"},
                                ],
                            }
                        ],
                    },
                    {"model_index": 2, "model_name": "sub", "status": "失败", "score": None, "comment": "", "error_message": "x", "latency_ms": 120, "sections": []},
                ],
            }
        ],
        summary={"批次ID": "batch-demo"},
    )

    out_path = tmp_path / "grade_result.xlsx"
    out_wb = load_workbook(out_path)

    # 模板法：不新建工作簿/工作表，保留模板中的 sheet 集合
    assert out_wb.sheetnames == ["成绩总览", "模型结果（长表）", "批改模型结果（宽表）", "维度汇总", "细则明细", "批次总览"]

    # 表格仍存在，且范围 ref 已扩展到包含“数据行 + 分隔空行”
    overview_tbl = out_wb["成绩总览"].tables["OverviewTbl"]
    assert overview_tbl.ref.endswith("3")
    model_long_tbl = out_wb["模型结果（长表）"].tables["ModelLong"]
    assert model_long_tbl.ref.endswith("4")

    # “查看模型/查看细则”使用 HYPERLINK 公式（模板法写公式，不写 hyperlink 对象）
    ws = out_wb["成绩总览"]
    headers = [c.value for c in ws[1]]
    view_model_col = headers.index("查看模型") + 1
    view_criteria_col = headers.index("查看细则") + 1
    assert str(ws.cell(row=2, column=view_model_col).value).startswith("=HYPERLINK(")
    assert str(ws.cell(row=2, column=view_criteria_col).value).startswith("=HYPERLINK(")
