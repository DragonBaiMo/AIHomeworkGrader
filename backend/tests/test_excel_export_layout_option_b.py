"""Excel 导出（选项B：模型结果宽表）回归测试。"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from app.util.excel_utils import ExcelExporter


def test_export_results_layout_option_b_sheets_and_headers(tmp_path: Path) -> None:
    exporter = ExcelExporter(tmp_path)
    exporter.export_results(
        [
            {
                "file_name": "张三.docx",
                "student_id": "2025001",
                "student_name": "张三",
                "score": 55.0,
                "score_rubric": 45.0,
                "score_rubric_max": 60.0,
                "status": "成功",
                "error_message": None,
                "comment": "总体评语",
                "detail_json": None,
                "aggregate_strategy": "mean",
                "grader_results": [
                    {
                        "model_index": 1,
                        "api_url": "http://a",
                        "model_name": "main",
                        "status": "成功",
                        "score": 55,
                        "comment": "c1",
                        "error_message": None,
                        "latency_ms": 100,
                        "sections": [
                            {
                                "name": "结构",
                                "max_score": 10,
                                "score": 8,
                                "comment": "结构不错",
                                "items": [
                                    {"name": "标题", "max_score": 5, "score": 4, "comment": "标题清晰"},
                                    {"name": "段落", "max_score": 5, "score": 4, "comment": "段落清楚"},
                                ],
                            }
                        ],
                    },
                    {
                        "model_index": 2,
                        "api_url": "http://b",
                        "model_name": "sub",
                        "status": "成功",
                        "score": 57,
                        "comment": "c2",
                        "error_message": None,
                        "latency_ms": 120,
                        "sections": [
                            {
                                "name": "结构",
                                "max_score": 10,
                                "score": 9,
                                "comment": "结构更好",
                                "items": [
                                    {"name": "标题", "max_score": 5, "score": 5, "comment": "很好"},
                                    {"name": "段落", "max_score": 5, "score": 4, "comment": "基本到位"},
                                ],
                            }
                        ],
                    },
                ],
            }
        ],
        summary={"批次ID": "batch-demo"},
        error_rows=[
            {
                "file_name": "李四.docx",
                "error_type": "解析校验错误",
                "error_message": "解析失败：" + ("很长的错误信息用于测试自动换行与行高自适应。" * 12),
            }
        ],
    )

    path = tmp_path / "grade_result.xlsx"
    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "Dashboard",
        "成绩总览",
        "模型结果（长表）",
        "批改模型结果",
        "维度汇总",
        "细则明细",
        "批次总览",
        "错误",
    ]

    # 生产环境增强：自动筛选 + Excel 表格（部分工作表）
    assert workbook["成绩总览"].auto_filter.ref is not None
    assert "TblOverview" in workbook["成绩总览"].tables
    assert workbook["批改模型结果"].auto_filter.ref is not None
    assert "TblModelsWide" in workbook["批改模型结果"].tables
    assert workbook["模型结果（长表）"].auto_filter.ref is not None
    assert "TblModelsLong" in workbook["模型结果（长表）"].tables
    assert workbook["错误"].auto_filter.ref is not None
    assert "TblErrors" in workbook["错误"].tables

    overview_headers = [cell.value for cell in workbook["成绩总览"][1]]
    assert "状态" in overview_headers
    assert "最终分（目标满分制）" in overview_headers
    assert "规则得分" in overview_headers
    assert "规则满分" in overview_headers
    assert "成功模型" in overview_headers
    assert "总体评语" in overview_headers
    # 对齐：数据区左对齐 + 顶端对齐
    overview_cell = workbook["成绩总览"]["E2"]
    assert overview_cell.alignment is not None
    assert overview_cell.alignment.horizontal == "left"
    assert overview_cell.alignment.vertical == "top"

    model_headers = [cell.value for cell in workbook["批改模型结果"][1]]
    assert "主模型名称" in model_headers
    assert "主模型状态" in model_headers
    assert "主模型评语" in model_headers
    assert "主模型耗时(ms)" in model_headers
    assert "副模型1名称" in model_headers
    assert "副模型1状态" in model_headers
    assert "副模型1评语" in model_headers
    assert "副模型2名称" not in model_headers
    assert "LLM总评语（主模型二次生成）" in model_headers

    dim_headers = [cell.value for cell in workbook["维度汇总"][1]]
    assert "汇总得分（成功模型均值）" in dim_headers
    assert "主模型得分" in dim_headers
    assert "副模型1得分" in dim_headers
    assert "副模型2得分" not in dim_headers
    assert "参与模型数" not in dim_headers

    criteria_headers = [cell.value for cell in workbook["细则明细"][1]]
    assert "细则名称" in criteria_headers
    assert "扣分原因/得分依据（取主模型）" in criteria_headers
    assert "参与模型数" not in criteria_headers

    error_headers = [cell.value for cell in workbook["错误"][1]]
    assert error_headers == ["文件名", "错误类型", "错误描述"]
    # 长文本应启用自动换行，并设置行高
    err_cell = workbook["错误"]["C2"]
    assert err_cell.alignment is not None
    assert err_cell.alignment.wrap_text is True
    assert (workbook["错误"].row_dimensions[2].height or 0) > 16

    # 独立异常清单也应具备筛选与表格
    exporter.export_errors([{"file_name": "李四.docx", "error_type": "解析校验错误", "error_message": "解析失败"}])
    err_path = tmp_path / "error_list.xlsx"
    err_wb = load_workbook(err_path)
    assert err_wb.sheetnames == ["异常清单"]
    err_ws = err_wb["异常清单"]
    assert err_ws.auto_filter.ref is not None
    assert "TblErrorList" in err_ws.tables
